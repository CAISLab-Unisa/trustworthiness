import pandas as pd
from openpyxl import load_workbook
import os

# path
trustability_file = r".\Trustability_Malign_Updated69.xlsx"
tsv_folder = r".\prediction"  # Folder with TSV files
output_file = r".\Trustability_NODR_UpdatedF.xlsx"

#  upload excel file 
wb = load_workbook(trustability_file)

# tsv file (in alphabetical order)
tsv_files = sorted([f for f in os.listdir(tsv_folder) if f.endswith(".tsv")])

# does not consider the last sheet (report)
sheets_to_update = wb.sheetnames[:-1]  

# check if the number of sheets is equal to the number of tsv files
if len(tsv_files) != len(sheets_to_update):
    raise ValueError(f"Errore: {len(tsv_files)} file TSV non corrispondono ai {len(sheets_to_update)} fogli nel file Excel!")

# scroll through the excel file
for sheet_name, tsv_file in zip(sheets_to_update, tsv_files):
    ws = wb[sheet_name]  # Seleziona il foglio corrente

    # 📖 read tsv file 
    tsv_path = os.path.join(tsv_folder, tsv_file)
    df_tsv = pd.read_csv(tsv_path, sep=",", engine="python")

    
    print(f"Elaborazione: {tsv_file} -> {sheet_name}")
    print(f"Dimensioni del file TSV: {df_tsv.shape}")

    #  check on column control
    if df_tsv.shape[1] < 4:
        raise ValueError(f"Il file {tsv_file} ha solo {df_tsv.shape[1]} colonne, non abbastanza per estrarre le colonne 2, 3 e 4.")

    #  extract only the columns I need
    df_selected = df_tsv.iloc[111:202, [1, 2, 3]].values    #cutting off two numbers to the first and one to the second
   
    # ✅ Verify that there are exactly 200 rows
   # if len(df_selected) != 361:
       # raise ValueError(f"Errore nel file {tsv_file}: sono state selezionate {len(df_selected)} righe invece di 361!")

    #  enter the values in the columns of the excel)
    for i, value_row in enumerate(df_selected, start=3):  # Inserisce esattamente dalla riga 2 a 201 (200 valori)
        ws[f"D{i}"] = value_row[0]  
        ws[f"E{i}"] = value_row[1]  
        ws[f"F{i}"] = value_row[2] 

# save
wb.save(output_file)

print(f" File salvato con successo: {output_file}")


